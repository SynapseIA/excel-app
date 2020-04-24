from flask import Flask, url_for, render_template, request, session
import openpyxl

app=Flask(__name__)

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/_pushData/', methods=['POST'])
def _pushData():
    input1=request.form.get('input1') or None
    input2=request.form.get('input2') or None
    input3=request.form.get('input3') or None
    input4=request.form.get('input4') or None

    if not input1 or not input2 or not input3 or not input4:
        return "Please provide all data <a href='/'> Push data <a>"
    else: 
        operation(input1, input2, input3, input4)
    return "done"


def operation(input1, input2, input3, input4):
    wb=openpyxl.load_workbook('excel-app.xlsx')
    Sheet1=wb.worksheets[0]
    Sheet1.cell(row=1, column=2).value=int(input1)
    Sheet1.cell(row=2, column=2).value=int(input2)
    Sheet1.cell(row=3, column=2).value=int(input3)
    Sheet1.cell(row=4, column=2).value=int(input4)
    wb.save('excel-app.xlsx')
    wb.close()

if __name__=="__main__":
    app.run(debug=False)